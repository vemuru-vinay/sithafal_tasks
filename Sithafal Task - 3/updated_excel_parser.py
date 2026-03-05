import os
import re
import pandas as pd
import openpyxl
from typing import Dict, Any, List, Tuple
from pprint import pprint

INPUT_FOLDER = "./testing_files"
FILE_NAME_STARTS_WITH = "" #set to test specific files
SUPPORTED_EXTENSIONS = (".xlsx", ".xls", ".csv")
OUTPUT_FOLDER = "./parsed_output"

HEADER_KEYWORDS = {
    "metric", "target", "consumed", "actual", "kpi", "value", "definition"
}

def is_foodrecord_meal_layout(df: pd.DataFrame) -> bool:
    """
    Detect Food/Fluid chart layout where meal sections are row titles such as
    Breakfast/Lunch and value marks are spread across multiple columns.
    """
    if df.empty or df.shape[1] < 3:
        return False

    text_df = df.astype(str).apply(lambda col: col.str.strip())
    meal_hits = 0
    has_comment = False
    has_pro_kj = False

    for _, row in text_df.iterrows():
        row_vals = [str(v).strip() for v in row.tolist() if str(v).strip() != ""]
        if not row_vals:
            continue
        for v in row_vals:
            low = re.sub(r"\s+", " ", v.lower())
            if re.match(r"^(breakfast|morning\s*tea|lunch|afternoon\s*tea|dinner|supper)\b", low):
                if "please record name of food/drink" in low:
                    meal_hits += 1
            if "comment" in low:
                has_comment = True
            if "pro" in low and "kj" in low:
                has_pro_kj = True

    return meal_hits >= 2 and (has_comment or has_pro_kj)

def extract_foodrecord_meta_rows(df: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    Capture known narrative/banner rows that must be retained as metadata in
    Food/Fluid meal sheets, including repeated sections.
    """
    if df.empty:
        return []

    out: List[Dict[str, Any]] = []
    text_df = df.astype(str).apply(lambda col: col.str.strip())

    def _match_meta_text(v: str) -> bool:
        low = re.sub(r"\s+", " ", str(v).strip().lower())
        if low == "":
            return False
        if "queensland health dietitian" in low:
            return True
        if "do not file" in low:
            return True
        if "food and fluid consumption chart" in low:
            return True
        if "food items in bold are high protein/energy" in low:
            return True
        if low.startswith("start date:") and "finish date:" in low and "commenced by:" in low:
            return True
        return False

    for i in range(len(text_df)):
        vals = [str(v).strip() for v in text_df.iloc[i].tolist() if str(v).strip() != ""]
        if not vals:
            continue
        if any(_match_meta_text(v) for v in vals):
            if len(vals) == 1:
                out.append({"rowIndex": i + 1, "Title": vals[0]})
            else:
                out.append({"rowIndex": i + 1, "content": vals})

    return out


def main():

    if not os.path.exists(OUTPUT_FOLDER):
        os.makedirs(OUTPUT_FOLDER)

    if not os.path.exists(INPUT_FOLDER):
        print(f"Error: Folder '{INPUT_FOLDER}' not found.")
        return

    files = [
        f for f in os.listdir(INPUT_FOLDER)
        if (f.lower().endswith(SUPPORTED_EXTENSIONS)
            and not f.startswith("~$")
            and f.startswith(FILE_NAME_STARTS_WITH))
    ]

    if not files:
        print("No supported files found.")
        return

    final_output = [] 

    for file in files:
        file_path = os.path.join(INPUT_FOLDER, file)
        sheets_dict = load_file(file_path)

        sheet_outputs = []  # stores each sheet result

        for sheet_name, df in sheets_dict.items():

            # Step 1: Split into vertical blocks
            initial_blocks, final_meta_data = split_blocks_and_context(df)

            all_clean_blocks = []

            # Step 2: Split horizontally + re-evaluate metadata
            for ib in initial_blocks:
                # Keep Food/Fluid meal layouts intact; vertical splitting separates
                # food-item labels from value columns and breaks meal parsing.
                if is_foodrecord_meal_layout(ib):
                    all_clean_blocks.append(ib)
                    continue

                vertical_chunks = split_block_vertically(ib)

                for chunk in vertical_chunks:
                    sub_blocks, sub_meta_data = split_blocks_and_context(chunk)

                    # combine metadata
                    final_meta_data.extend(sub_meta_data)

                    # accumulate blocks
                    for blk in sub_blocks:
                        all_clean_blocks.append(blk)

            # Step 3: Clean each block
            cleaned_tables = []
            for blk in all_clean_blocks:
                cleaned = clean_block(blk)
                if isinstance(cleaned, list):
                    cleaned_tables.extend(cleaned)
                else:
                    cleaned_tables.append(cleaned)
            cleaned_tables = merge_continuation_tables(cleaned_tables)

            # Backfill known Food/Fluid narrative rows that can be dropped when
            # meal tables are extracted from mixed blocks.
            if is_foodrecord_meal_layout(df):
                existing_rows = {int(m.get("rowIndex")) for m in final_meta_data if isinstance(m, dict) and "rowIndex" in m}
                for m in extract_foodrecord_meta_rows(df):
                    if int(m["rowIndex"]) not in existing_rows:
                        final_meta_data.append(m)
                        existing_rows.add(int(m["rowIndex"]))
                final_meta_data.sort(key=lambda x: x.get("rowIndex", 10**9) if isinstance(x, dict) else 10**9)

            # Step 4: Add sheet result
            sheet_outputs.append({
                "sheet_name": sheet_name,
                "meta_data": final_meta_data,
                "tables": cleaned_tables
            })

        # Step 5: Add final file entry
        final_output.append({
            "file_name": file,
            "sheets": sheet_outputs
        })

        # Write JSON to file
        output_path = os.path.join(OUTPUT_FOLDER, file.replace(".xlsx", ".json").replace(".xls", ".json").replace(".csv", ".json"))
        
        with open(output_path, "w", encoding="utf-8") as f:
            import json
            json.dump(final_output[-1], f, indent=4, ensure_ascii=False)
            
        print(f"Saved JSON -> {output_path}")

    pprint(final_output)

def merge_continuation_tables(tables: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Merge tables that are actually continuations of the previous table.
    Handles headerless continuation blocks:

    """
    def _looks_like_bed_data_header(headers: List[str]) -> bool:
        if not headers:
            return False
        first = str(headers[0]).strip().lower()
        return bool(re.fullmatch(r"bed\s*\d+", first))

    merged: List[Dict[str, Any]] = []
    for table in tables:
        if not merged:
            merged.append(table)
            continue

        prev = merged[-1]
        prev_headers = prev.get("headers", [])
        cur_headers = table.get("headers", [])
        prev_named = bool(str(prev.get("table_name", "")).strip())
        cur_named = bool(str(table.get("table_name", "")).strip())

        # Keep explicitly named section tables separate.
        if prev_named or cur_named:
            merged.append(table)
            continue

        # Case 1: exact same schema -> append rows
        if prev_headers == cur_headers:
            prev["rows"].extend(table.get("rows", []))
            continue

        # Case 2: current table looks like headerless continuation data
        if (
            len(prev_headers) == len(cur_headers)
            and _looks_like_bed_data_header(cur_headers)
            and re.fullmatch(r"bed\s*\d+", str(prev_headers[0]).strip().lower()) is None
        ):
            # Convert current table "headers" into the first data row.
            first_row = {prev_headers[i]: cur_headers[i] for i in range(len(prev_headers))}

            # Re-map existing current rows by position to previous headers.
            remapped_rows = []
            for row in table.get("rows", []):
                vals = [row.get(h, "") for h in cur_headers]
                remapped_rows.append({prev_headers[i]: vals[i] for i in range(len(prev_headers))})

            prev["rows"].append(first_row)
            prev["rows"].extend(remapped_rows)
            continue

        merged.append(table)

    return merged

def load_file(filepath: str) -> Dict[str, pd.DataFrame]:
    """Loads file into a dictionary of DataFrames (one per sheet)."""
    if filepath.endswith(".csv"):
        return {"CSV_Data": pd.read_csv(filepath, header=None).fillna("")}
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws = unmerge_and_fill(ws)
        rows = [row for row in ws.iter_rows(values_only=True)]
        df = pd.DataFrame(rows).fillna("")
        if not df.empty:
            sheets_data[sheet_name] = df
            
    return sheets_data

def unmerge_and_fill(ws: openpyxl.worksheet.worksheet.Worksheet) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Detects merged cells in an Excel worksheet, unmerges them, and fills
    the previously merged range with the value from the top-left cell.
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = cell_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(cell_range))
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = top_left_value
    return ws

def split_blocks_and_context(df: pd.DataFrame) -> Tuple[List[pd.DataFrame], List[Dict[str, Any]]]:
    """Splits a DataFrame into table blocks + extracted meta data rows."""
    is_empty = df.astype(str).apply(lambda x: x.str.strip()).eq("").all(axis=1)
    blocks = []
    meta_data_rows = []
    current_block_indices = []

    def _to_clean_values(row: pd.Series) -> List[str]:
        vals = []
        for v in row.tolist():
            t = str(v).strip()
            if t != "" and t.lower() != "nan":
                vals.append(t)
        return vals

    def _infer_value_type(values: List[str]) -> str:
        if not values:
            return "empty"
        joined = " ".join(values).strip()
        lowered = joined.lower()
        if any(v.upper() in {"RED", "YELLOW", "GREEN", "AMBER"} for v in values):
            return "status"
        if joined.endswith("?"):
            return "question"
        if re.fullmatch(r"[-+]?\d+(\.\d+)?%?", joined):
            return "number"
        if re.fullmatch(r"\d+\s*/\s*\d+", joined):
            return "ratio"
        if re.match(r"^\d{1,2}:\d{2}$", joined):
            return "time"
        if re.search(r"\b\d{1,2}:\d{2}\b", lowered):
            return "datetime_or_time"
        if re.fullmatch(r"\d+[.)]\s+.+", joined):
            return "list_item"
        return "text"

    def _build_meta_entry(row_idx: int, row: pd.Series) -> Dict[str, Any]:
        values = _to_clean_values(row)
        # Export 1-based row numbering to match spreadsheet row labels.
        entry: Dict[str, Any] = {"rowIndex": row_idx + 1, "content": values}
        if not values:
            return entry

        # Single-cell narrative rows are treated as titles in output.
        # Keep this before inline-pair parsing so rows like "UNIT: ICU-West"
        # are emitted as Title, not inline_pairs.
        if (
            len(values) == 1
            and not re.fullmatch(r"\d+[.)]\s+.+", values[0].strip())
            and not values[0].strip().endswith(":")
        ):
            return {"rowIndex": row_idx + 1, "Title": values[0]}

        first = values[0]
        if first.endswith(":"):
            key = first[:-1].strip()
            value_parts = values[1:]
            entry["key"] = key
            if len(value_parts) == 1:
                entry["value"] = value_parts[0]
            else:
                entry["values"] = value_parts
                entry["value"] = " ".join(value_parts).strip()
            entry["value_type"] = _infer_value_type(value_parts)
        else:
            inline_pairs = []
            for token in values:
                m = re.match(r"^\s*([^:]+?)\s*:\s*(.+)\s*$", token)
                if not m:
                    continue
                k = m.group(1).strip()
                v = m.group(2).strip()
                inline_pairs.append({
                    "key": k,
                    "value": v,
                    "value_type": _infer_value_type([v])
                })
            if inline_pairs:
                entry["inline_pairs"] = inline_pairs
        return entry

    def _build_column_pair_meta(indices: List[int], chunk_df: pd.DataFrame) -> Dict[str, Any]:
        top_vals = _to_clean_values(chunk_df.iloc[0])
        bottom_vals = _to_clean_values(chunk_df.iloc[1])
        pair_count = min(len(top_vals), len(bottom_vals))
        pairs = []
        for i in range(pair_count):
            raw_key = top_vals[i]
            row_value = bottom_vals[i]
            m = re.match(r"^\s*([^:]+?)\s*:\s*(.+)\s*$", raw_key)
            if m:
                key = m.group(1).strip()
                inline_v = m.group(2).strip()
                value = f"{inline_v} | {row_value}" if row_value else inline_v
            else:
                key = raw_key.strip()
                value = row_value
            if key == "":
                continue
            pairs.append({
                "key": key,
                "value": value,
                "value_type": _infer_value_type([value])
            })

        return {
            "rowIndex": indices[0] + 1,
            "rowIndex_end": indices[-1] + 1,
            "content": top_vals + bottom_vals,
            "column_pairs": pairs
        }

    def is_key_value_context(chunk_df: pd.DataFrame) -> bool:
        """
        Detect key/value context blocks and keep them out of table parsing.
        This accepts multi-column value tails as long as each row starts with
        a content header token ending with ':'.
        """
        normalized = chunk_df.copy()
        normalized = normalized.where(pd.notna(normalized), "")
        normalized = normalized.astype(str).apply(lambda col: col.str.strip())
        if normalized.shape[0] < 1:
            return False

        for _, row in normalized.iterrows():
            row_vals = [v for v in row.tolist() if v != ""]
            if not row_vals:
                continue
            if len(row_vals) < 2:
                return False
            if not str(row_vals[0]).strip().endswith(":"):
                return False
        return True

    def is_two_row_column_context(chunk_df: pd.DataFrame) -> bool:
        """
        Detect two-row narrative metadata laid out as columns, e.g.
        Row 1: Nurse Station | Charge: Sarah J. | Pager: 5590
        Row 2: Overflow      | 1. Pending ...    | 2. Empty
        """
        normalized = chunk_df.copy()
        normalized = normalized.where(pd.notna(normalized), "")
        normalized = normalized.astype(str).apply(lambda col: col.str.strip())
        if normalized.shape[0] != 2:
            return False

        non_empty_cols = [col for col in normalized.columns if not normalized[col].eq("").all()]
        if len(non_empty_cols) < 2:
            return False

        r0 = [v for v in normalized.iloc[0][non_empty_cols].tolist() if v != ""]
        r1 = [v for v in normalized.iloc[1][non_empty_cols].tolist() if v != ""]
        if min(len(r0), len(r1)) < 2:
            return False

        has_inline_colon = any(re.match(r"^\s*[^:]+:\s*\S+", v) for v in r0)
        has_numbered_value = any(re.match(r"^\s*\d+[.)]\s+\S+", v) for v in r1)
        same_width = len(r0) == len(r1)
        return has_inline_colon and (has_numbered_value or same_width)

    def process_chunk(indices):
        if not indices: return
        if len(indices) < 2:
            row_idx = indices[0]
            meta_data_rows.append(_build_meta_entry(row_idx, df.loc[row_idx]))
            return

        chunk_df = df.loc[indices]
        non_empty_cols = [c for c in chunk_df.columns if not chunk_df[c].astype(str).str.strip().eq("").all()]

        def _is_bed_header_row(series: pd.Series) -> bool:
            vals = [str(v).strip() for v in series.tolist()]
            bed_hits = sum(1 for v in vals if re.fullmatch(r"bed\s*\d+", v, flags=re.IGNORECASE))
            return bed_hits >= 3

        def _is_axis_alias_row(series: pd.Series) -> bool:
            vals = [str(v).strip() for v in series.tolist()]
            if not vals:
                return False
            # Expect empty top-left and letter aliases across columns: A, B, C...
            if vals[0] != "":
                return False
            alias_hits = sum(1 for v in vals[1:] if re.fullmatch(r"[A-Za-z]", v))
            return alias_hits >= 3

        def _row_contains_meal_title(series: pd.Series) -> bool:
            vals = [str(v).strip() for v in series.tolist() if str(v).strip() != ""]
            for v in vals:
                low = re.sub(r"\s+", " ", v.lower())
                if re.match(r"^(breakfast|morning\s*tea|lunch|afternoon\s*tea|dinner|supper)\b", low):
                    if "please record name of food/drink" in low:
                        return True
            return False

        # Narrative blocks (single effective column) should remain context.
        if len(non_empty_cols) == 1:
            for row_idx in indices:
                meta_data_rows.append(_build_meta_entry(row_idx, df.loc[row_idx]))
            return

        # Sparse form banners/titles (few rows, very few filled cells) are context.
        sparse_rows = []
        for _, row in chunk_df.iterrows():
            row_vals = [str(v).strip() for v in row.tolist() if str(v).strip() != ""]
            sparse_rows.append(len(row_vals))
        if len(indices) <= 4 and sparse_rows and max(sparse_rows) <= 2:
            for row_idx in indices:
                meta_data_rows.append(_build_meta_entry(row_idx, df.loc[row_idx]))
            return

        # Also treat repeated merged banners (same token duplicated across columns)
        # as context instead of a tiny noisy table.
        if len(indices) <= 4:
            banner_like = True
            for _, row in chunk_df.iterrows():
                row_vals = [str(v).strip() for v in row.tolist() if str(v).strip() != ""]
                if not row_vals:
                    continue
                freq: Dict[str, int] = {}
                for v in row_vals:
                    freq[v] = freq.get(v, 0) + 1
                uniq_count = len(freq)
                max_rep = max(freq.values())
                if uniq_count > 2 or max_rep < 3:
                    banner_like = False
                    break
            if banner_like:
                for row_idx in indices:
                    meta_data_rows.append(_build_meta_entry(row_idx, df.loc[row_idx]))
                return

        # Keep Bed matrix layouts as table blocks (don't strip row as context title).
        if len(indices) >= 2:
            row0 = chunk_df.iloc[0]
            row1 = chunk_df.iloc[1]
            if _is_bed_header_row(row0):
                if str(row1.iloc[0]).strip() != "":
                    blocks.append(chunk_df)
                    return

        if is_key_value_context(chunk_df):
            for row_idx in indices:
                meta_data_rows.append(_build_meta_entry(row_idx, df.loc[row_idx]))
            return

        if is_two_row_column_context(chunk_df):
            meta_data_rows.append(_build_column_pair_meta(indices, chunk_df))
            return

        row0 = chunk_df.iloc[0]
        row1 = chunk_df.iloc[1]

        # Keep meal title row inside the block for food-record extraction.
        if _row_contains_meal_title(row0):
            blocks.append(chunk_df)
            return

        # Keep pure alias-axis row (A/B/C...) with following row-label matrix as table.
        # This prevents that axis row from being exported as meta_data content.
        if _is_axis_alias_row(row0) and str(row1.iloc[0]).strip() != "":
            blocks.append(chunk_df)
            return

        count_0 = (row0 != "").sum()
        count_1 = (row1 != "").sum()

        is_sparser_title = (count_0 < count_1) and (count_1 > 1)
        is_single_title = (count_0 == 1) and (count_1 > 1)

        if is_sparser_title or is_single_title:
            meta_data_rows.append(_build_meta_entry(indices[0], df.loc[indices[0]]))
            remaining_indices = indices[1:]
            if len(remaining_indices) < 2:
                meta_data_rows.append(_build_meta_entry(remaining_indices[0], df.loc[remaining_indices[0]]))
            else:
                blocks.append(df.loc[remaining_indices])
        else:
            blocks.append(chunk_df)

    for idx, empty in is_empty.items():
        if not empty:
            current_block_indices.append(idx)
        else:
            if current_block_indices:
                process_chunk(current_block_indices)
                current_block_indices = []

    if current_block_indices:
        process_chunk(current_block_indices)

    return blocks, meta_data_rows

def split_block_vertically(df: pd.DataFrame) -> List[pd.DataFrame]:
    """
    Takes a DataFrame (a block of rows) and checks if it can be 
    split into multiple side-by-side tables based on empty columns.
    """
    # 1. Identify which columns are completely empty in this specific block
    is_col_empty = (df == "").all(axis=0)
    
    # If no empty columns, return the block as is
    if not is_col_empty.any():
        return [df]

    sub_blocks = []
    start_col = 0
    total_cols = df.shape[1]
    
    # 2. Iterate through columns to find split points
    # We use numerical indexing because column names might be duplicates or messy
    for col_idx in range(total_cols):
        # Check if this column index corresponds to an empty column
        # (Using .iloc to access the boolean series by position)
        if is_col_empty.iloc[col_idx]:
            # If we found an empty column, slice the data before it
            if col_idx > start_col:
                sub_blocks.append(df.iloc[:, start_col:col_idx])
            
            # Move start pointer to the next column
            start_col = col_idx + 1
            
    # 3. Don't forget the last chunk after the final empty column
    if start_col < total_cols:
        sub_blocks.append(df.iloc[:, start_col:])
        
    return sub_blocks

def clean_block(df_block: pd.DataFrame) -> Any:
    """
    Clean a detected block:
    - Detect if table is vertical -> rotate to horizontal
    - Detect header row for horizontal tables
    - Return structured JSON: headers + rows
    """

    was_rotated = False
    df = df_block.copy().reset_index(drop=True)
    df = df.where(pd.notna(df), "")
    df = df.loc[:, ~(df == "").all(axis=0)]

    def _normalize_header_name(h: str) -> str:
        t = str(h).strip()
        if t.lower() == "bay":
            return "Bed"
        return t

    def _excel_col_name(idx: int) -> str:
        """
        Convert 0-based index to Excel-like column name:
        0 -> A, 25 -> Z, 26 -> AA
        """
        name = ""
        n = idx
        while True:
            n, rem = divmod(n, 26)
            name = chr(ord("A") + rem) + name
            if n == 0:
                break
            n -= 1
        return name

    def _is_bed_matrix_layout(df_in: pd.DataFrame) -> bool:
        if len(df_in) < 3 or df_in.shape[1] < 3:
            return False
        row0 = [str(v).strip() for v in df_in.iloc[0].tolist()]
        row1 = [str(v).strip() for v in df_in.iloc[1].tolist()]
        bed_hits = sum(1 for v in row0 if re.fullmatch(r"bed\s*\d+", v, flags=re.IGNORECASE))
        has_row_label = row1[0] != ""
        alias_hits = sum(1 for v in row1[1:] if re.fullmatch(r"[A-Za-z]", v))
        return bed_hits >= 3 and has_row_label and alias_hits >= 3

    def _looks_like_alias_grid_without_top_header(df_in: pd.DataFrame) -> bool:
        """
        Detect layout where bed headers are missing and first row is data like:
        Name | A | B | C ...
        Formula | ...
        Rate    | ...
        """
        if len(df_in) < 3 or df_in.shape[1] < 3:
            return False
        row0 = [str(v).strip() for v in df_in.iloc[0].tolist()]
        if row0[0] == "":
            return False
        if re.fullmatch(r"bed\s*\d+", row0[0], flags=re.IGNORECASE):
            return False
        alias_hits = sum(1 for v in row0[1:] if re.fullmatch(r"[A-Za-z]", v))
        if alias_hits < 3:
            return False

        # Ensure this behaves like a matrix: first column has labels on next rows.
        next_labels = [str(v).strip() for v in df_in.iloc[1:, 0].tolist()]
        label_hits = sum(1 for v in next_labels if v != "")
        return label_hits >= 2

    def _extract_foodrecord_vertical_tables(df_in: pd.DataFrame) -> List[Dict[str, Any]]:
        """
        Parse Food/Fluid chart sections where meal titles are vertical row labels and
        there is no clean horizontal header row for value columns.
        """
        if df_in.empty or df_in.shape[1] < 3:
            return []

        text_df = df_in.astype(str).apply(lambda col: col.str.strip())

        def _is_meal_title(text: str) -> bool:
            t = str(text).strip()
            if t == "":
                return False
            low = re.sub(r"\s+", " ", t.lower())
            has_meal_word = re.match(
                r"^(breakfast|morning\s*tea|lunch|afternoon\s*tea|dinner|supper)\b",
                low
            ) is not None
            has_prompt = "please record name of food/drink" in low
            return has_meal_word and has_prompt

        def _clean_title(text: str) -> str:
            t = str(text).strip()
            t = re.sub(r"\(.*?please circle.*?\)", "", t, flags=re.IGNORECASE)
            t = re.sub(r"please\s+record\s+name\s+of\s+food/drink", "", t, flags=re.IGNORECASE)
            t = re.sub(r"\s+", " ", t).strip(" -:\t")
            return t

        # Detect the vertical title column (typically the "food item" column).
        title_col = -1
        best_hits = 0
        for c in text_df.columns:
            hits = sum(1 for v in text_df[c].tolist() if _is_meal_title(v))
            if hits > best_hits:
                best_hits = hits
                title_col = c
        if title_col == -1 or best_hits < 2:
            return []

        # Meal title rows can be merged across many columns, so detect them by
        # low unique-value cardinality instead of raw non-empty count.
        title_rows = []
        for i in range(len(text_df)):
            cell = text_df.iloc[i][title_col]
            if not _is_meal_title(cell):
                continue
            row_vals = [str(v).strip() for v in text_df.iloc[i].tolist() if str(v).strip() != ""]
            if not row_vals:
                continue
            normalized = [re.sub(r"\s+", " ", v.lower()) for v in row_vals]
            uniq = set(normalized)
            if len(uniq) <= 3 and any(_is_meal_title(v) for v in row_vals):
                title_rows.append(i)
        if not title_rows:
            return []

        forced_comment_col = None
        forced_pro_col = None
        for c in text_df.columns:
            if c == title_col:
                continue
            col_vals = [str(v).strip().lower() for v in text_df[c].tolist()]
            if forced_comment_col is None and any("comment" in v for v in col_vals):
                forced_comment_col = c
            if forced_pro_col is None and any(("pro" in v and "kj" in v) for v in col_vals):
                forced_pro_col = c

        all_tables: List[Dict[str, Any]] = []
        for idx, start in enumerate(title_rows):
            end = title_rows[idx + 1] if idx + 1 < len(title_rows) else len(text_df)
            title_text = _clean_title(text_df.iloc[start][title_col])
            if title_text == "":
                title_text = "Table"

            section_rows = []
            for r in range(start + 1, end):
                row = text_df.iloc[r]
                item = str(row[title_col]).strip()
                joined_row = " ".join(v for v in row.tolist() if str(v).strip() != "").lower()
                if "food items in bold" in joined_row or joined_row.startswith("start date:"):
                    break
                if item == "":
                    # A blank in the food item axis ends this meal section.
                    break
                if _is_meal_title(item):
                    break
                section_rows.append(r)

            if not section_rows:
                continue

            # Keep columns that actually carry section data.
            value_cols = []
            for c in text_df.columns:
                if c == title_col:
                    continue
                has_data = any(str(text_df.iloc[r][c]).strip() != "" for r in section_rows)
                label_hint = str(text_df.iloc[start][c]).strip().lower()
                is_comment_or_pro = ("comment" in label_hint) or ("pro" in label_hint and "kj" in label_hint)
                if has_data or is_comment_or_pro:
                    value_cols.append(c)
            for forced_col in [forced_comment_col, forced_pro_col]:
                if forced_col is not None and forced_col not in value_cols:
                    value_cols.append(forced_col)

            if not value_cols:
                continue

            # Preserve Comment / Pro kJ labels when discoverable, otherwise use A/B/C...
            col_name_map: Dict[int, str] = {}
            letter_idx = 0
            for c in value_cols:
                if forced_comment_col is not None and c == forced_comment_col:
                    col_name_map[c] = "Comment"
                    continue
                if forced_pro_col is not None and c == forced_pro_col:
                    col_name_map[c] = "Pro kJ"
                    continue
                label_hint = str(text_df.iloc[start][c]).strip().lower()
                if "comment" in label_hint:
                    col_name_map[c] = "Comment"
                    continue
                if "pro" in label_hint and "kj" in label_hint:
                    col_name_map[c] = "Pro kJ"
                    continue
                col_name_map[c] = _excel_col_name(letter_idx)
                letter_idx += 1

            headers = ["food item"] + [col_name_map[c] for c in value_cols]
            out_rows = []
            for r in section_rows:
                row = text_df.iloc[r]
                record = {"food item": str(row[title_col]).strip()}
                for c in value_cols:
                    record[col_name_map[c]] = str(row[c]).strip()
                out_rows.append(record)

            all_tables.append({
                "table_name": title_text,
                "headers": headers,
                "rows": out_rows,
                "was_rotated": False,
                "table_orientation": "horizontal"
            })

        return all_tables

    # Special-case: meal sections from Food and Fluid Consumption Chart.
    meal_tables = _extract_foodrecord_vertical_tables(df)
    if meal_tables:
        return meal_tables

    # Special-case: header row with Bed labels + alias row (A/B/C...).
    if _is_bed_matrix_layout(df):
        top = [str(v).strip() for v in df.iloc[0].tolist()]
        headers = ["row_label"]
        headers.extend([v for v in top[1:] if v != ""])
        headers = [_normalize_header_name(h) for h in headers]

        # Keep the alias row (e.g., Name | A..F) as the first data row.
        data = df.iloc[1:].reset_index(drop=True)
        data = data.iloc[:, :len(headers)]
        data.columns = headers
        return {
            "headers": headers,
            "rows": data.astype(str).to_dict(orient="records"),
            "was_rotated": False,
            "table_orientation": "horizontal"
        }

    # Special-case: no bed header row, first row is alias/value row (Name | A..F).
    if _looks_like_alias_grid_without_top_header(df):
        row0 = [str(v).strip() for v in df.iloc[0].tolist()]
        value_cols = df.shape[1] - 1

        headers = ["Row Label"]
        for idx in range(value_cols):
            alias = row0[idx + 1] if idx + 1 < len(row0) else ""
            if alias == "":
                alias = chr(ord("A") + idx)
            headers.append(alias)
        headers = [_normalize_header_name(h) for h in headers]

        data = df.iloc[:, :len(headers)].reset_index(drop=True)
        data.columns = headers
        return {
            "headers": headers,
            "rows": data.astype(str).to_dict(orient="records"),
            "was_rotated": False,
            "table_orientation": "horizontal"
        }

    def _value_like(text: str) -> bool:
        t = str(text).strip().lower()
        if t == "":
            return True
        if re.fullmatch(r"[-+]?\d+(\.\d+)?%?", t):
            return True
        if re.fullmatch(r"\d+\s*/\s*\d+", t):
            return True
        if " out of " in t:
            return True
        if "min" in t:
            return True
        return False

    def _header_like(text: str) -> bool:
        t = str(text).strip().lower()
        if "result value" in t:
            return True
        return any(re.search(rf"\b{re.escape(k)}\b", t) for k in HEADER_KEYWORDS)

    def _metric_block(df_in: pd.DataFrame) -> bool:
        non_empty_cols = [c for c in df_in.columns if not df_in[c].astype(str).str.strip().eq("").all()]
        if len(non_empty_cols) < 2:
            return False

        first_row_vals = [
            str(v).strip() for v in df_in.iloc[0][non_empty_cols].tolist() if str(v).strip() != ""
        ]
        if any(_header_like(v) for v in first_row_vals):
            return False

        sample = df_in[non_empty_cols].astype(str).apply(lambda col: col.str.strip())
        rows_with_metric = sample[sample.iloc[:, 0] != ""]
        if rows_with_metric.empty:
            return False

        second_col = rows_with_metric.iloc[:, 1].tolist()
        value_like_ratio = sum(1 for v in second_col if _value_like(v)) / len(second_col)
        return value_like_ratio >= 0.7

    # KPI list blocks (metric + value [+notes]) should not use first row as header.
    if _metric_block(df):
        non_empty_cols = [c for c in df.columns if not df[c].astype(str).str.strip().eq("").all()]
        core = df[non_empty_cols].astype(str).apply(lambda col: col.str.strip())

        headers = ["Metric", "Value"]
        out_rows = []
        for _, row in core.iterrows():
            metric = row.iloc[0] if len(row) > 0 else ""
            value = row.iloc[1] if len(row) > 1 else ""
            if len(row) > 2 and row.iloc[2] != "":
                headers = ["Metric", "Value", "Notes"]
                out_rows.append({"Metric": metric, "Value": value, "Notes": row.iloc[2]})
            else:
                out_rows.append({"Metric": metric, "Value": value})
        headers = [_normalize_header_name(h) for h in headers]

        return {
            "headers": headers,
            "rows": out_rows,
            "was_rotated": True,
            "table_orientation": "vertical"
        }

    # 1. DETECT VERTICAL TABLES
    # Guardrail: only attempt vertical rotation for 2-column blocks.
    non_empty_col_count = (~(df == "").all(axis=0)).sum()
    vinfo = detect_vertical_headers(df)
    if vinfo["is_vertical"] and non_empty_col_count == 2:
        df = rotate_vertical_table(df, vinfo)
        df = df.reset_index(drop=True)
        was_rotated = True  

    # 2. Detect header row
    header_idx = detect_header_row(df)

    # Guardrail: if the detected header is the last row, parsing would produce
    # an empty table. In that case, treat the block as row-label matrix data.
    if header_idx >= len(df) - 1 and df.shape[1] >= 2:
        first_col = df.iloc[:, 0].astype(str).str.strip()
        non_empty_first = (first_col != "").sum()
        if non_empty_first >= 2:
            value_cols = df.shape[1] - 1
            headers = ["Row Label"] + [_excel_col_name(i) for i in range(value_cols)]
            headers = [_normalize_header_name(h) for h in headers]

            data = df.iloc[:, :len(headers)].reset_index(drop=True)
            data.columns = headers
            return {
                "headers": headers,
                "rows": data.astype(str).to_dict(orient="records"),
                "was_rotated": was_rotated,
                "table_orientation": "horizontal"
            }

    # 3. Merge multi-row headers
    header = merge_multirow_header(df, header_idx)
    header = [_normalize_header_name(h) for h in header]

    # 3. REMOVE HEADER ROW FROM DATA
    data = df.drop(index=list(range(header_idx + 1))).reset_index(drop=True)

    # 4. CLEAN EMPTY COLUMNS
    data = data.loc[:, ~(data == "").all(axis=0)]
    if len(header) != data.shape[1]:
        header = header[:data.shape[1]]
        if len(header) < data.shape[1]:
            header.extend([f"col_{i}" for i in range(len(header), data.shape[1])])

    # Cleanup: matrix-style tables with A/B/C... columns should expose
    # a readable first header instead of autogenerated col_0.
    if (
        header
        and re.fullmatch(r"col_\d+", str(header[0]).strip(), flags=re.IGNORECASE)
        and len(header) >= 2
    ):
        alias_hits = sum(1 for h in header[1:] if re.fullmatch(r"[A-Za-z]+", str(h).strip()))
        if alias_hits >= 3:
            header[0] = "Row Label"

    data.columns = header  # assign headers

    # 5. CONVERT TO JSON-FRIENDLY FORMAT
    return {
        "headers": header,
        "rows": data.astype(str).to_dict(orient="records"),
        "was_rotated": was_rotated,
        "table_orientation": "vertical" if was_rotated else "horizontal"
    }


def detect_vertical_headers(df: pd.DataFrame) -> Dict[str, Any]:
    """Legacy function to detect transposed tables. Preserved for future use."""
    if df.shape[1] < 2: return {"is_vertical": False}

    # Only consider vertical detection for true 2-column blocks.
    non_empty_cols = [c for c in df.columns if not df[c].fillna("").astype(str).str.strip().eq("").all()]
    if len(non_empty_cols) != 2:
        return {"is_vertical": False}

    col0 = df.iloc[:, 0]
    valid_indices = [i for i, v in enumerate(col0) if pd.notna(v) and str(v).strip() != ""]
    if len(valid_indices) < 2: return {"is_vertical": False}
    
    first_idx = valid_indices[0]
    first_val_neighbor = df.iloc[first_idx, 1]
    if pd.isna(first_val_neighbor) or str(first_val_neighbor).strip() == "":
        if len(valid_indices) > 1: valid_indices = valid_indices[1:]

    col0_vals = [str(col0[i]).strip() for i in valid_indices]
    unique_ratio = len(set(col0_vals)) / len(col0_vals)
    numeric_ratio = sum(1 for v in col0_vals if re.match(r"^\d+$", v)) / len(col0_vals)

    if unique_ratio > 0.9 and numeric_ratio < 0.5:
        return {"is_vertical": True, "start_index": valid_indices[0], "headers": col0_vals}

    return {"is_vertical": False}

def rotate_vertical_table(df: pd.DataFrame, vinfo: Dict[str, Any]) -> pd.DataFrame:
    """
    Rotates a vertical table into a horizontal table.
    Example vertical:
        Metric
        ED LWBS%
        High Acuity
        LOS

        Value
        0.02
        15%
        300

    After rotation:
        Metric        Value
        ED LWBS%      0.02
        High Acuity   15%
        LOS           300
    """

    # vinfo['headers'] = list of vertical header labels
    headers = vinfo["headers"]

    # STEP 1: Collect all columns after vertical headers
    # DataFrame has multiple columns, but vertical tables typically use col0 and col1
    columns = df.shape[1]

    # Extract col0 and col1 as arrays
    col0 = df.iloc[:, 0].tolist()
    col1 = df.iloc[:, 1].tolist() if columns > 1 else [""] * len(col0)

    # STEP 2: Split into header list and value list
    # vinfo['headers'] tells the names that appear vertically
    header_count = len(headers)

    # Now construct horizontal header
    horizontal_header = ["Metric", "Value"]  # simple case

    # STEP 3: Build data rows by pairing (col0[i], col1[i])
    data_start = vinfo["start_index"] + header_count
    data_rows = []

    for i in range(header_count):
        metric_name = headers[i]
        metric_value = col1[vinfo["start_index"] + i] if (vinfo["start_index"] + i) < len(col1) else ""
        data_rows.append([metric_name, metric_value])

    return pd.DataFrame([horizontal_header] + data_rows)

def merge_multirow_header(df: pd.DataFrame, header_idx: int) -> List[str]:
    """
    Merge multiple header rows (e.g., row0 + row1) into a single header list.
    Rules:
      - Stop merging when a row has fewer filled cells than the row below
      - Empty cells inherit the value from above
      - Final header is cleaned and flattened
    """

    # Collect rows that belong to header zone
    header_zone = []
    prev_non_empty = None

    for i in range(header_idx + 1):  
        row = df.iloc[i].fillna("").astype(str).str.strip().tolist()

        non_empty_count = sum([1 for v in row if v != ""])
        if prev_non_empty is not None and non_empty_count < prev_non_empty:
            break  # header ended

        header_zone.append(row)
        prev_non_empty = non_empty_count

    # Now merge header rows column-wise
    merged = []
    num_cols = len(header_zone[0])

    for c in range(num_cols):
        parts = []
        for r in header_zone:
            if r[c] != "":
                parts.append(r[c])
        merged.append(" ".join(parts).strip())

    # Clean fallback names
    merged = [h if h != "" else f"col_{i}" for i, h in enumerate(merged)]
    return merged


def detect_header_row(df: pd.DataFrame, max_row_idx: int = 5) -> int:
    """Legacy function to detect header row index. Preserved for future use."""
    scores = {}
    scan_depth = min(max_row_idx + 1, len(df))

    for i in range(scan_depth):
        row = df.iloc[i]
        values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() != ""]
        if not values: continue
        string_count = len(values)
        unique_count = len(set(values))
        score = (unique_count * 3) + string_count
        score -= i * 0.5
        scores[i] = score

    return max(scores, key=scores.get) if scores else 0


if __name__ == "__main__":
    main()
