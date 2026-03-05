import os
import re
import pandas as pd
import openpyxl
from typing import Dict, Any, List, Tuple
from pprint import pprint

# --------------------------------------------------
# CONFIG
# --------------------------------------------------

INPUT_FOLDER = "./testing_files"
FILE_NAME_STARTS_WITH = "" #set to test specific files
SUPPORTED_EXTENSIONS = (".xlsx", ".xls", ".csv")

# --------------------------------------------------
# CORE HELPERS
# --------------------------------------------------

def unmerge_and_fill(ws: openpyxl.worksheet.worksheet.Worksheet) -> openpyxl.worksheet.worksheet.Worksheet:
    """
    Detects merged cells in an Excel worksheet, unmerges them, and fills
    the previously merged range with the value from the top-left cell.
    """
    merged_ranges = list(ws.merged_cells.ranges)
    for cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = cell_range.bounds
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(cell_range))
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = top_left_value
    return ws


def detect_header_row(df: pd.DataFrame, max_row_idx: int = 5) -> int:
    """Legacy function to detect header row index. Preserved for future use."""
    scores = {}
    scan_depth = min(max_row_idx + 1, len(df))

    for i in range(scan_depth):
        row = df.iloc[i]
        values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip() != ""]
        if not values: continue
        string_count = len(values)
        unique_count = len(set(values))
        score = (unique_count * 3) + string_count
        score -= i * 0.5
        scores[i] = score

    return max(scores, key=scores.get) if scores else 0


def detect_vertical_headers(df: pd.DataFrame) -> Dict[str, Any]:
    """Legacy function to detect transposed tables. Preserved for future use."""
    if df.shape[1] < 2: return {"is_vertical": False}
    col0 = df.iloc[:, 0]
    valid_indices = [i for i, v in enumerate(col0) if pd.notna(v) and str(v).strip() != ""]
    if len(valid_indices) < 2: return {"is_vertical": False}
    
    first_idx = valid_indices[0]
    first_val_neighbor = df.iloc[first_idx, 1]
    if pd.isna(first_val_neighbor) or str(first_val_neighbor).strip() == "":
        if len(valid_indices) > 1: valid_indices = valid_indices[1:]

    col0_vals = [str(col0[i]).strip() for i in valid_indices]
    unique_ratio = len(set(col0_vals)) / len(col0_vals)
    numeric_ratio = sum(1 for v in col0_vals if re.match(r"^\d+$", v)) / len(col0_vals)

    if unique_ratio > 0.9 and numeric_ratio < 0.5:
        return {"is_vertical": True, "start_index": valid_indices[0], "headers": col0_vals}

    return {"is_vertical": False}


def process_block(df_block: pd.DataFrame) -> List[Dict[str, Any]]:
    """
    1. Copies the block.
    2. Drops columns that are completely empty (all cells are "").
    3. Renumbers columns (0, 1, 2...) so they look clean.
    4. Returns a preview dictionary.
    """
    # Create a copy and fill NaNs to ensure clean string manipulation
    data_df = df_block.copy().reset_index(drop=True).fillna("")

    # --- NEW: Drop Empty Columns ---
    # 1. Identify columns where EVERY row is empty string ""
    #    (We use .eq("") to compare and .all() to check if ALL rows match)
    empty_cols = [col for col in data_df.columns if data_df[col].eq("").all()]
    
    # 2. Drop them
    data_df.drop(columns=empty_cols, inplace=True)

    # 3. Renumber columns to start from 0 again (e.g., if we drop 0, Col 1 becomes 0)
    data_df.columns = range(data_df.shape[1])
    # -------------------------------

    # Return the first 5 rows as a list of dictionaries
    # #return data_df.head(5).astype(str).to_dict(orient="records")

    # Alternatively, print everything
    return data_df.astype(str).to_dict(orient="records")


# --------------------------------------------------
# FILE PROCESSING
# --------------------------------------------------

def load_file(filepath: str) -> Dict[str, pd.DataFrame]:
    """Loads file into a dictionary of DataFrames (one per sheet)."""
    if filepath.endswith(".csv"):
        return {"CSV_Data": pd.read_csv(filepath, header=None).fillna("")}
    
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheets_data = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws = unmerge_and_fill(ws)
        rows = [row for row in ws.iter_rows(values_only=True)]
        df = pd.DataFrame(rows).fillna("")
        if not df.empty:
            sheets_data[sheet_name] = df
            
    return sheets_data


def split_blocks_and_context(df: pd.DataFrame) -> Tuple[List[pd.DataFrame], List[Dict[str, Any]]]:
    """Splits a DataFrame into 'Blocks' (tables) and 'Context' (titles/metadata)."""
    is_empty = df.astype(str).apply(lambda x: x.str.strip()).eq("").all(axis=1)
    blocks = []
    context_rows = []
    current_block_indices = []

    def process_chunk(indices):
        if not indices: return
        if len(indices) < 2:
            row_idx = indices[0]
            context_rows.append({"rowIndex": row_idx, "content": df.loc[row_idx][df.loc[row_idx] != ""].astype(str).tolist()})
            return

        chunk_df = df.loc[indices]
        row0 = chunk_df.iloc[0]
        row1 = chunk_df.iloc[1]
        count_0 = (row0 != "").sum()
        count_1 = (row1 != "").sum()

        is_sparser_title = (count_0 < count_1) and (count_1 > 1)
        is_single_title = (count_0 == 1) and (count_1 > 1)

        if is_sparser_title or is_single_title:
            context_rows.append({"rowIndex": indices[0], "content": row0[row0 != ""].astype(str).tolist()})
            remaining_indices = indices[1:]
            if len(remaining_indices) < 2:
                context_rows.append({"rowIndex": remaining_indices[0], "content": df.loc[remaining_indices[0]][df.loc[remaining_indices[0]] != ""].astype(str).tolist()})
            else:
                blocks.append(df.loc[remaining_indices])
        else:
            blocks.append(chunk_df)

    for idx, empty in is_empty.items():
        if not empty:
            current_block_indices.append(idx)
        else:
            if current_block_indices:
                process_chunk(current_block_indices)
                current_block_indices = []

    if current_block_indices:
        process_chunk(current_block_indices)

    return blocks, context_rows

def split_block_vertically(df: pd.DataFrame) -> List[pd.DataFrame]:
    """
    Takes a DataFrame (a block of rows) and checks if it can be 
    split into multiple side-by-side tables based on empty columns.
    """
    # 1. Identify which columns are completely empty in this specific block
    is_col_empty = (df == "").all(axis=0)
    
    # If no empty columns, return the block as is
    if not is_col_empty.any():
        return [df]

    sub_blocks = []
    start_col = 0
    total_cols = df.shape[1]
    
    # 2. Iterate through columns to find split points
    # We use numerical indexing because column names might be duplicates or messy
    for col_idx in range(total_cols):
        # Check if this column index corresponds to an empty column
        # (Using .iloc to access the boolean series by position)
        if is_col_empty.iloc[col_idx]:
            # If we found an empty column, slice the data before it
            if col_idx > start_col:
                sub_blocks.append(df.iloc[:, start_col:col_idx])
            
            # Move start pointer to the next column
            start_col = col_idx + 1
            
    # 3. Don't forget the last chunk after the final empty column
    if start_col < total_cols:
        sub_blocks.append(df.iloc[:, start_col:])
        
    return sub_blocks


# --------------------------------------------------
# MAIN
# --------------------------------------------------

def main():
    if not os.path.exists(INPUT_FOLDER):
        print(f"Error: Folder '{INPUT_FOLDER}' not found.")
        return

    files = [f for f in os.listdir(INPUT_FOLDER) if (f.lower().endswith(SUPPORTED_EXTENSIONS) and not f.startswith("~$") and f.startswith(FILE_NAME_STARTS_WITH))]

    if not files:
        print("No supported files found.")
        return

    for file in files:
        print("\n" + "=" * 80)
        print(f"📄 FILE: {file}")
        print("=" * 80)

        sheets_dict = load_file(os.path.join(INPUT_FOLDER, file))

        for sheet_name, df in sheets_dict.items():
            print(f"\n   >>> SHEET: {sheet_name}")
            print("   " + "-" * 40)

            # --- PASS 1: Initial Vertical Split ---
            # Separates main data chunks from top/bottom context
            initial_blocks, final_context = split_blocks_and_context(df)

            final_blocks = []

            # --- PASS 2: Horizontal Split & Re-Check ---
            for ib in initial_blocks:
                # 1. Split vertically stacked blocks horizontally (e.g. Table | Sidebar)
                vertical_chunks = split_block_vertically(ib)
                
                for chunk in vertical_chunks:
                    # 2. RECURSIVE STEP:
                    # Pass the horizontally split chunk BACK into split_blocks_and_context.
                    # This allows sidebars (which were previously "shadowed" by the main table)
                    # to be evaluated on their own. The scanner will now see the empty rows
                    # inside the sidebar and correctly break it into Context rows.
                    sub_blocks, sub_context = split_blocks_and_context(chunk)
                    
                    final_blocks.extend(sub_blocks)
                    final_context.extend(sub_context)

            # Sort context by row index so it prints in reading order
            final_context.sort(key=lambda x: x['rowIndex'])

            if final_context:
                print("\n   CONTEXT (Non-Tabular Data)")
                for c in final_context:
                    print(f"   Row {c['rowIndex']}: {c['content']}")
            
            print(f"\n   Blocks detected: {len(final_blocks)}\n")

            for i, block in enumerate(final_blocks):
                print(f"   --- BLOCK {i + 1} Preview ---")
                pprint(process_block(block))
                print()

if __name__ == "__main__":
    main()


